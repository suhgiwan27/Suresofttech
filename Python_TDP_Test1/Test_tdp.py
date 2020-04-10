# -*- coding: utf-8 -*

import openpyxl
import string
import tkinter.filedialog
import tkinter as tk


# 기능사양분석서(FSA)
FSAfinder = tk.Tk()
FSAfinder.withdraw()
FSApath = tk.filedialog.askopenfilename(filetypes=[('Microsoft Excel File(.xlsx)', '.xlsx'), ('All files', '*')],
                                        title='Open FSA File',
                                        initialfile='*.xlsx')
fuc_wb = openpyxl.load_workbook(FSApath)
fuc_sheet = fuc_wb.active

# TDP
resultwb = openpyxl.Workbook()
resultsheet = resultwb.active

# TDP 포맷 만들기(첫째 라인)
resultsheet['B1'] = "No"
resultsheet['C1'] = "Test Requirement"
resultsheet['D1'] = ""
resultsheet['E1'] = "TestCase"
resultsheet['F1'] = ""
resultsheet['G1'] = "Test Scenario"
resultsheet['H1'] = ""
resultsheet['M1'] = ""
resultsheet["N1"] = "비고"
resultsheet["O1"] = "Signal"

# TDP 포맷 만들기(둘째 라인)
resultsheet['B2'] = ""
resultsheet['C2'] = "Req_Num"
resultsheet['D2'] = "Description"
resultsheet['E2'] = "TC Num"
resultsheet['F2'] = "Description"
resultsheet['G2'] = "구분"
resultsheet['H2'] = "Description"
resultsheet['I2'] = "time"
resultsheet['J2'] = "Function"
resultsheet['K2'] = "variable"
resultsheet['L2'] = "연산자"
resultsheet['M2'] = "예상값"
resultsheet['N2'] = ""
resultsheet['O2'] = ""

# TDP 포맷 병합 //가로
resultsheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
resultsheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
resultsheet.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)
resultsheet.merge_cells(start_row=1, start_column=12, end_row=1, end_column=13)

# TDP 포맷 병합 //세로
resultsheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
resultsheet.merge_cells(start_row=1, start_column=14, end_row=2, end_column=14)
resultsheet.merge_cells(start_row=1, start_column=15, end_row=2, end_column=15)

# 기능사양 분석서 포맷 가운데 정렬
for absorption in range(2, 16):
    for absorption_row in range(1,3):
        resultsheet.cell(row=absorption_row, column=absorption).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

#변수 정리
req_col = 9
req_row = 4                       #기능분석사양서/ Req ID 4행
req_row2 = req_row -1             #기능분석사양서/ Req_ID에 해당하는 ccp, can 신호 count해주는 변수
req_count =1                      #기능분석사양서/ Req_ID 변수
req_Numcount = 1                  #기능분석사양서/ Req_ID의 text 뽑아내기 위한 변수

count = 1                         #기능분석사양서/ Req_ID에 해당하는 행(row) 갯수

req_text = ""                     #기능분석사양서/ Req_ID Text 불러오기.
req_text2 = ""                    #기능분석사양서/ Req_row2 에 해당되는 Text를 저장
req_text3 = ""                    #기능분석사양서/ List var[]에 담겨있는 값들을 text 한줄로 저장하기 위한변수

tdp_text = ""                     #기능분석사양서/ ccp, can 신호 등을 tdp_text 변수에 한줄로 저장
Signal_row = 3                    #TDP/ Signal Row 변수
Signal_row2 = 3
Tc_Num_row = 3                    #TDP/ Tc_Num Row 변수
Tc_Num_count = 1                  #TDP/ Req_id 하나당 해당되는 Tc01, Tc02 등의 count

tdp_Tcnum = "TC_"
tdp_Req = "Req_"
Tdp_modul = ""

# req 행 갯수 구하기
for a in range(0, fuc_sheet.max_row):
    req_text = fuc_sheet.cell(column= 7, row= req_row).value
    if req_text == None:
        req_row +=1
        count +=1
    else:
        print("")
        print(req_count)
        req_count +=1
        list_var = []

        # 옆으로 넘어가서 텍스트 인식하고 실행
        for b in range(0, count):                ## for문 첫시작
            req_text2 = fuc_sheet.cell(column=11, row=req_row2).value
            if(req_text2 !=None):
                tdp_text += req_text2 + '\n'

            req_row2 = req_row2 +1
        req_row2 = req_row2 -count                ## 한로직 종료

        for c in range(0, count):
            req_text2 = fuc_sheet.cell(column=12, row=req_row2).value
            if(req_text2 !=None):
                tdp_text += req_text2 + '\n'

            req_row2 = req_row2 +1
        req_row2 = req_row2 -count

        for d in range(0, count):
            req_text2 = fuc_sheet.cell(column=13, row=req_row2).value
            if(req_text2 !=None):
                tdp_text += req_text2 +'\n'

            req_row2 = req_row2 +1
        req_row2 = req_row2 -count

        for e in range(0, count):
            req_text2 = fuc_sheet.cell(column=15, row=req_row2).value
            if(req_text2 !=None):
                tdp_text += req_text2 + '\n'

            req_row2 = req_row2 +1

#Text에서 '\n' 인식해서 각각 리스트에 저장해줌
        list_var = tdp_text.splitlines()
        print(list_var)
        list_var = list(set(list_var))
        n = len(list_var)
        print(list_var)

#첫째줄에 적기위하여 Req_text3에 리스트에 저장된 Text들을 Req_text3에 저장
        for aa in range(0, n):
            print(list_var[aa])
            req_text3 += list_var[aa] + '\n'

#TDP/ Req_Num, Signal에 Text 기입
        resultsheet.cell(column=2, row=Signal_row).value = req_Numcount
        resultsheet.cell(column=15, row=Signal_row).value = req_text3

#TDP/ Signal 한줄씩 기입 ########################################################################
        for Signal_loop in range(0, count):
            # 음영 색 지정
            Signal_Color = openpyxl.styles.PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            resultsheet.cell(column=15, row=Signal_row2).fill = Signal_Color
            for bb in range(0, n):
                resultsheet.cell(column=15, row=Signal_row2).value = list_var[bb]
                Signal_row2 += 1



#기능분석사양서/ Module Location의 Text 불러오기
        Tdp_modul = fuc_sheet.cell(column=5, row= req_row2 - count).value

        if(Tdp_modul ==None):
            Tdp_modul = "미정"
        else:
            Tdp_modul = Tdp_modul.split('/', 1)[0]
            print(Tdp_modul)

#TDP/ Req_Num에 Text 기입
        resultsheet.cell(column=3, row=Signal_row).value = fuc_sheet.cell(column=7, row=req_row2 - count).value

#TDP/ Signal row 병합
        resultsheet.merge_cells(start_row=Signal_row, start_column=2, end_row=Signal_row + (count * n) - 1, end_column=2)
        resultsheet.merge_cells(start_row=Signal_row, start_column=3, end_row=Signal_row + (count * n) - 1, end_column=3)
        #resultsheet.merge_cells(start_row=Signal_row, start_column=15, end_row=Signal_row + (count * n) - 1, end_column=15)


#TDP/ TC_Num에 Text 기입
        for Tc_loop in range (0, count):
            resultsheet.merge_cells(start_row=Tc_Num_row, start_column=5, end_row=Tc_Num_row + n - 1, end_column=5)
            resultsheet.cell(column= 5, row=Tc_Num_row).value = tdp_Tcnum + str(Tdp_modul) + "_" + str(format(Tc_Num_count,'03'))
            Tc_Num_count +=1
            Tc_Num_row +=n

# 값 초기화
        Signal_row = Signal_row + (count * n)    # tc 갯수만큼 합침
        count = 1
        req_row = req_row + 1
        tdp_text = ""
        req_text3 =""
        req_Numcount += 1

# 셀 크기 조정
resultsheet.column_dimensions['O'].width = 25

# 텍스트 줄바꿈 적용, 중앙 정렬
for rows in resultsheet.iter_rows(min_row=resultsheet.min_row, max_row=resultsheet.max_row,
                                  min_col=resultsheet.min_column, max_col=resultsheet.max_column):
    for cell in rows:
        cell.alignment = openpyxl.styles.Alignment(wrapText=True, horizontal='center', vertical='center')

# 셀 너비 자동 조정
for i in list(string.ascii_uppercase):
    resultsheet.column_dimensions[i].bestFit = True

# 음영 색 지정
Color_Green = openpyxl.styles.PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
Color_Pastel_Blue = openpyxl.styles.PatternFill(start_color='8FAADC', end_color='8FAADC', fill_type='solid')
Color_Light_Blue = openpyxl.styles.PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type='solid')
Color_Pastel_Yellow = openpyxl.styles.PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
Color_Pastel_Orange = openpyxl.styles.PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid')
Color_Red = openpyxl.styles.PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
for i in range(resultsheet.min_row, resultsheet.min_row + 2):
    resultsheet.cell(row=i, column=resultsheet.min_column).fill = Color_Green
    resultsheet.cell(row=i, column=resultsheet.min_column + 1).fill = Color_Pastel_Blue
    resultsheet.cell(row=i, column=resultsheet.min_column + 2).fill = Color_Pastel_Blue
    resultsheet.cell(row=i, column=resultsheet.min_column + 3).fill = Color_Light_Blue
    resultsheet.cell(row=i, column=resultsheet.min_column + 4).fill = Color_Light_Blue
    for j in range(5, 10):
        resultsheet.cell(row=i, column=resultsheet.min_column + j).fill = Color_Pastel_Yellow
    resultsheet.cell(row=i, column=resultsheet.min_column + 10).fill = Color_Pastel_Orange
    resultsheet.cell(row=i, column=resultsheet.min_column + 11).fill = Color_Pastel_Orange
    resultsheet.cell(row=i, column=resultsheet.min_column + 12).fill = Color_Red

# 테두리 자동작성
for j in range(resultsheet.min_column, resultsheet.max_column + 1):
    for i in range(resultsheet.min_row, resultsheet.max_row + 1):
        border_var = openpyxl.styles.Border(
            left=openpyxl.styles.Side(border_style="thin", color='FF000000'),
            right=openpyxl.styles.Side(border_style="thin", color='FF000000'),
            top=openpyxl.styles.Side(border_style="thin", color='FF000000'),
            bottom=openpyxl.styles.Side(border_style="thin", color='FF000000'),
            diagonal=openpyxl.styles.Side(border_style="thin", color='FF000000'),
            diagonal_direction=0,
            outline=openpyxl.styles.Side(border_style="thin", color='FF000000'),
            vertical=openpyxl.styles.Side(border_style="thin", color='FF000000'),
            horizontal=openpyxl.styles.Side(border_style="thin", color='FF000000')
        )
        resultsheet.cell(column=j, row=i).border = border_var

#엑셀 파일 저장
TDPsaver = tk.Tk()
TDPsaver.withdraw()
TDPpath = tkinter.filedialog.asksaveasfilename(filetypes=[('Microsoft Excel File(.xlsx)', '.xlsx'), ('All files', '*')],
                                               title='Save as..',
                                               initialfile='*.xlsx')

resultwb.save(TDPpath)